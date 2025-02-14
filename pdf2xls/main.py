import camelot
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Путь к вашему PDF-файлу
pdf_file = "acc.pdf"

# Пробуем режим stream
try:
    tables = camelot.read_pdf(pdf_file, pages='all', flavor='stream')
    print(f"Найдено таблиц: {len(tables)}")
except Exception as e:
    print("Ошибка при чтении PDF:", e)

excel_file = "output.xlsx"

# Записываем каждую найденную таблицу в отдельный лист Excel
with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
    for i, table in enumerate(tables):
        df = table.df  # таблица в виде DataFrame
        sheet_name = f"Table_{i+1}"
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"Таблица {i+1} записана в лист {sheet_name}")

# Загружаем сохранённый Excel-файл для добавления Excel-таблиц
wb = load_workbook(excel_file)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Определяем диапазон ячеек с данными (предполагается, что первая строка — заголовки)
    max_row = ws.max_row
    max_column = ws.max_column
    end_column = get_column_letter(max_column)
    table_range = f"A1:{end_column}{max_row}"
    
    # Создаем Excel-таблицу (ListObject)
    table_obj = Table(displayName=f"Table_{sheet_name}", ref=table_range)
    
    # Задаем стиль для таблицы (можно менять стиль по вкусу)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table_obj.tableStyleInfo = style
    ws.add_table(table_obj)
    print(f"Excel Table добавлена для листа {sheet_name}")

wb.save(excel_file)
print("Конвертация завершена. Файл сохранен как", excel_file)
